# Johnathan Mai
# Function: Analyzes multiple parts lists

'''
Directory of Functions

combine_parts_lists(*args)
	Purpose: Currently combines all files in a folder (presumably .xlsx parts lists) into one sheet for easy editing.
	Input: optional string, path to folder. If not, prompts you to select folder.
	Returns: Nothing, creates master.xlsx with all parts list at destination.

add_pl_to_file(xls_path, master_path)
	Purpose: Takes the Parts List from the xls_path and adds the important data (Part Name, QTY, Part ID) to a combined .xlsx file.
	Input: xls_path(string)- path to parts list (.xlsx)
			master_path(string)- path to master .xlsx file to append to.
	Returns: Nothing, creates master .xlsx file if not already existing.

get_paths_from_folder(folder_path)
	Purpose: Gets paths to all files in a folder. 
	Input: string, path to folder
	Returns: list of strings, all of the paths in a list

make_folder(folder_name)
	Purpose: Makes a folder at current operating path with specified name.
	Input: folder_name(string)- desired name of folder.
	Returns: Nothing

'''

# Python Imports
import os
from tkinter import *
from tkinter import filedialog
import numpy as np
from os import path

# Non-Python Imports
from pandas import read_excel
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Functions
def combine_parts_lists(*args):
	if not args:
		print('Choose your folder')
		try:
			root = Tk()
			root.withdraw()
			filePath = filedialog.askdirectory()
			xls_paths = get_paths_from_folder(filePath)
		except:
			print('No File Selected.')	
	elif len(args) == 1 and type(args[0]) == str:	
		try:
			xls_paths = get_paths_from_folder(args[0])
		except Exception as e:
			print(e.args[-1])
	else:
		sys.exit('Pass the folder path or leave arguments empty.')

	# Grabbing Part Name, QTY, Part ID columns from all .xlsx files in folder
	master_file_path = make_folder('Combined_Parts_List') + '\\' + 'Combined_Parts_List.xlsx'
	for file in xls_paths:
		add_pl_to_file(file, master_file_path)
	return
'''
	This function currently combines all part lists in the specified folder into one excel for easy editing. 
	Can't really programatically get coupons until a more consistent naming scheme is used. 
		i.e. T4277: the 21 parts on the bottom also have letter C in part ID AND BA #

	So for now it gathers the parts names, ids, and QTY into one xls, then manually select the coupons vs. parts.
'''

def count_parts():
	print('count_parts')

def count_coupons():
	print('count_coupons')

def add_pl_to_file(xls_path, master_path):
	dfs = read_excel(xls_path, engine='openpyxl')
	# Find row that houses column names
	index_qty = pd.DataFrame([ dfs[col].astype(str).str.contains('QTY', na=False) for col in dfs ]).transpose()
	dfs.columns = dfs.loc[index_qty.any(axis=1)].loc[1].values.flatten().tolist() # Sets new columns because current template has NaN as first row in xls which leads to weird columns

	# Part Name, QTY, Part ID targeted
	output_df = pd.DataFrame([dfs['Part Name'], dfs['QTY'], dfs['Part ID']]).transpose()
	output_df.columns = ['Part Name', 'QTY', 'Part ID']

	# Getting Traveler Number from XLS Path
	slash_index = [pos for pos, char in enumerate(xls_path) if char == '\\']
	sheet_name = xls_path[slash_index[len(slash_index)-1]+1:(len(xls_path)-5)]

	# Writes Sheet to Master File
	print('Added parts list for: ' + sheet_name)
	if not os.path.exists(master_path):
		wb = Workbook()
		wb.save(master_path)
		writer = pd.ExcelWriter(master_path, engine='xlsxwriter')
		output_df.to_excel(writer, sheet_name=sheet_name, index=False)
		writer.save()
	else:
		wb = load_workbook(master_path)
		writer = pd.ExcelWriter(master_path, mode='a', engine='openpyxl')
		writer.book = wb
		writer.sheets = dict( (ws.title, ws) for ws in wb.worksheets )
		output_df.to_excel(writer, sheet_name=sheet_name, index=False)
		writer.save()
	return

def get_paths_from_folder(folder_path):
	paths = []
	for path in os.listdir(folder_path):
		paths.append(folder_path + '\\' + path)
	return paths

def make_folder(folder_name):
	folder_path = os.getcwd() + '\\' + folder_name
	if not os.path.isdir(folder_path):
		os.mkdir(folder_path)
	else: 
		print('Folder exists already')
	return folder_path

def main():
	print('main()')

if __name__ == '__main__':
	main()