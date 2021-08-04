# Johnathan Mai
# Requires Python 3
# Install pandas library: (pip install pandas) in command line

## Imports ##
# Python Modules
import tkinter as tkint
from tkinter import filedialog
from datetime import datetime as dt
import datetime
import os
import sys

# Non-Python Modules
from pandas import read_excel
import pandas
from openpyxl import load_workbook
from openpyxl import Workbook

## Functions ##

def remove_colon(strang):
	return strang.split(':')[-1]

def remove_equal(strang):
	return strang.split('=')[-1]

def time_difference(time_start, time_end):
	# Function: Gets the time difference between two datetime STRINGS. (HH:MM:SS)
	# Returns: Time difference in seconds

	t_start = dt.strptime(time_start, '%H:%M:%S')
	t_end = dt.strptime(time_end, '%H:%M:%S')
	t_elapsed = t_start - t_end
	if t_elapsed.total_seconds() < 0: # Special case for when time goes into the next day since I don't include AM/PM
		t_start = t_start + datetime.timedelta(hours=13)
		t_end = t_end + datetime.timedelta(hours=1)
		t_elapsed = t_start - t_end
	return t_elapsed.total_seconds()

def get_file_name(filePath):
	# Function: Gets name of file from the end of path.
	# Returns: Name of file (no extension). If extension longer than 3 letters, change the 4
	slash_index = [pos for pos, char in enumerate(filePath) if char == '\\']
	return filePath[slash_index[len(slash_index)-1]+1:(len(filePath)-5)]

def get_paths_from_folder(folderPath):
	# Function: Returns paths of all files [list] in a folder
	# Returns: ^
	paths = []
	for path in os.listdir(folderPath):
		paths.append(folderPath + '\\' + path)
	return paths

def parse_log(filePath):
	# Function: Parses the log at the file path into the data fields listed below.
	# Returns: Excel sheet in designated folder.

	# File Open
	try:
		printerlog = open(filePath, 'r')
	except:
		sys.exit('No File Selected')
		
	# Parse
	data_fields = ['L', 'D.Tm', 'B', 'F', 'FL', 'FR', 
	'CoaterSp', 'BlowerF', 'RPMBlowerSp', 'setp', 'real', 'Tlaser', 
	'P', 'O2', 'Tbuild', 'Par', 'Pca', 'dPfil'] 
	# DATAFIELDS ONLY WORK FOR TRAVELER NUMBERS T4213 AND BEYOND (OCTOBER 12, 2020) 
	# BECAUSE 'BlowerSp' IS DIFFERENT VARIABLE BEFORE THAT

	data_fields_cutoff1 = dt.strptime('10/11/2020', '%m/%d/%Y') 
	data_fields_cutoff2 = dt.strptime('2/7/2020', '%m/%d/%Y')
	corrector = 0 # Corrector for earlier printer logs that had a different format.

	log_date = get_date_from_log(filePath)
	if data_fields_cutoff2 < log_date < data_fields_cutoff1:
		data_fields[8] = 'ModBlowerSp'
	elif log_date < data_fields_cutoff2:
		data_fields.remove('RPMBlowerSp')
		data_fields.remove('setp')
		data_fields.remove('real')
		corrector = 3

	rows = []
	final_rows = []

	# Checks lines in log file if it contains the data_fields to identify which lines to pull from
	for temp_line in printerlog:
		if all([value in temp_line for value in data_fields]): 
			rows.append(temp_line[:-1])

	# MANUAL SPLICING EDIT THIS IF LOG FORMAT CHANGES ~~~
	data_fields = ['Date', 'Time', 'AM/PM'] + data_fields
	data_fields.insert(5, 'Accumulated Time (s)')
	if log_date > data_fields_cutoff1:
		data_fields.insert(18, 'O2 ppm')
		data_fields.insert(19, 'O2 %1')
	accumulated_time = 0

	# Messy parsing I know SORRY
	for row in rows:
		corrector = 0
		if log_date < data_fields_cutoff2:
			corrector = 3
		temp_split = row.split()
		temp_row = []

		temp_row.append(temp_split[0]) # Date
		temp_row.append(temp_split[1]) # Time
		temp_row.append(temp_split[2]) # AM or PM
		temp_row.append(temp_split[4]) # L
		temp_row.append(temp_split[6]) # D.Tm
		if len(final_rows) < 1: # Accumulated Time
			temp_row.append(accumulated_time)
		else:
			accumulated_time = accumulated_time + time_difference(temp_split[1], final_rows[len(final_rows)-1][1])
			temp_row.append(accumulated_time)
		temp_row.append(remove_colon(temp_split[7])) # B
		temp_row.append(temp_split[9]) # F
		temp_row.append(temp_split[11]) # FL
		temp_row.append(temp_split[13]) # FR
		temp_row.append(remove_colon(temp_split[14])) # CoaterSp
		temp_row.append(temp_split[16]) # Blower F
		if corrector == 0:
			temp_row.append('') # RPMBlowerSp
			temp_row.append(remove_equal(temp_split[18])) # setp
			temp_row.append(remove_equal(temp_split[19])) # real
		temp_row.append(temp_split[21-corrector]) # Tlaser
		temp_row.append(remove_colon(temp_split[22-corrector])) # P
		temp_row.append(remove_colon(temp_split[23-corrector])) # O2
		if log_date > data_fields_cutoff1:
			temp_row.append(temp_split[25]) # O2 ppm
			temp_row.append(temp_split[28]) # O2 %1
		else:
			corrector += 4
		temp_row.append(remove_colon(temp_split[30-corrector])) # Tbuild
		temp_row.append(remove_colon(temp_split[31-corrector])) # Par
		temp_row.append(remove_colon(temp_split[32-corrector])) # Pca
		temp_row.append(remove_colon(temp_split[33-corrector])) # dPfil


		final_rows.append(temp_row)

	# Write to xls
	xlsFolderPath = os.getcwd() + '\\' + 'ParsedLogs'
	if not os.path.isdir(xlsFolderPath):
		os.mkdir(xlsFolderPath)
	excelPath = xlsFolderPath + '\\' + get_file_name(filePath)
	writer = pandas.ExcelWriter(excelPath + '.xlsx', engine='xlsxwriter')

	df = pandas.DataFrame(final_rows)
	df.columns = data_fields
	df.to_excel(writer, sheet_name=get_machine_from_log(filePath), index=False)
	writer.save()

	# Close
	printerlog.close()

def get_date_from_log(filePath):
	# Function: Gets the date/time of first layer print of log file, through same process of parse_log().
	# Returns: python datetime object including date and time of day.

	try:
		printerlog = open(filePath, 'r')
	except:
		sys.exit('No File Selected.')
	row = ''
	for temp_line in printerlog:
		if 'Logging started' in temp_line:
			row = temp_line
			break
	row_split = row.split()[4] + ' ' + row.split()[5] + row.split()[6] + ' ' + row.split()[8] + row.split()[9]
	return dt.strptime(row_split, '%B %d,%Y %I:%M%p')

def get_machine_from_log(filePath):
	# Function: Determines what machine produced the log file.
	# Returns: String, machine name that produced log

	try:
		printerlog = open(filePath, 'r')
	except:
		sys.exit('No File Selected.')
	for temp_line in printerlog:
		if 'Machine: ' in temp_line:
			machine = temp_line.split()[-1]
			break
		else: 
			continue
	return machine

def compile_data_from_logs(folder_path, master_path):

	file_paths = get_paths_from_folder(folder_path)
	o2_data = {}
	master_df = pandas.DataFrame()
	for path in file_paths:
		temp_log = read_excel(path, engine='openpyxl')
		temp_wb = load_workbook(path)
		temp_machine = temp_wb.sheetnames[0]
		if '47' in temp_machine:
			print(temp_machine)
			temp_name = pandas.Series([get_file_name(path)])
			temp_data = temp_log['O2'].transpose()
			temp_df = pandas.DataFrame(temp_name.append(temp_data, ignore_index=True)).transpose()
			master_df = master_df.append(temp_df)
			print(master_df)
			
	master_wb = Workbook()
	master_wb.save(master_path)
	writer = pandas.ExcelWriter(master_path, engine='xlsxwriter')
	master_df.to_excel(writer, sheet_name='O2 Data', index=False)
	writer.save()
'''
	else:
		master_wb = load_workbook(master_path)
		writer = pandas.ExcelWriter(master_path, mode='a', engine='openpyxl')
		writer.book = master_wb
		writer.sheets = dict( (ws.title, ws) for ws in master_wb.worksheets )

		# just need to append after last row now

		master_ws = master_wb.active
		startrow = master_ws.max_row
		temp_df.to_excel(writer, index=False)
		writer.save()
'''

def main():
	try:
		root = tkint.Tk()
		root.withdraw()
		filePath = filedialog.askopenfilename()
		parse_log(filePath)
	except:
		print('No File Selected.')

if __name__ == '__main__':
	main()

print('Imported parseLogs!')